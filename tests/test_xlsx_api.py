from io import BytesIO
import os
import unittest

from fastapi.testclient import TestClient
from openpyxl import load_workbook

os.environ["DISABLE_TG_BOOT"] = "1"

from main import app


class XlsxApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(app)

    def test_build_xlsx_success(self):
        payload = {
            "file_name": "Balances_03.04.2026_11-00.xlsx",
            "sheets": [
                {
                    "name": "Operator Alert",
                    "rows": [
                        {
                            "merchant": "A",
                            "currency": "INR",
                            "balance": 1000,
                            "Operators_Net change_Last_72_hours - Operator_name → Δ_avg_daily_usd": 12345.67,
                        },
                        {
                            "merchant": "B",
                            "currency": "USD",
                            "balance": 2000,
                            "status": "ok",
                            "Operators_Net change_Last_72_hours - Operator_name → Δ_avg_daily_usd": None,
                        },
                    ],
                },
                {
                    "name": "Merchant Alert",
                    "rows": [],
                },
            ],
        }

        response = self.client.post("/xlsx/build", json=payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "Balances_03.04.2026_11-00.xlsx",
            response.headers["content-disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Operator Alert", "Merchant Alert"])

        ws = workbook["Operator Alert"]
        self.assertEqual(
            [cell.value for cell in ws[1]],
            ["merchant", "currency", "balance", "Δ_avg_daily_usd", "status"],
        )
        self.assertEqual([cell.value for cell in ws[2]], ["A", "INR", 1000, 12345.67, None])
        self.assertEqual([cell.value for cell in ws[3]], ["B", "USD", 2000, None, "ok"])
        self.assertEqual(ws["C2"].number_format, "#,##0")
        self.assertEqual(ws["D2"].number_format, "#,##0")
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertEqual(ws.auto_filter.ref, "A1:E3")

        empty_ws = workbook["Merchant Alert"]
        self.assertEqual(empty_ws.max_row, 1)
        self.assertEqual(empty_ws.max_column, 1)
        self.assertIsNone(empty_ws["A1"].value)

    def test_build_xlsx_invalid_payload(self):
        response = self.client.post(
            "/xlsx/build",
            json={
                "file_name": "broken.xlsx",
                "sheets": "not-an-array",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("sheets", response.json()["detail"])


if __name__ == "__main__":
    unittest.main()
