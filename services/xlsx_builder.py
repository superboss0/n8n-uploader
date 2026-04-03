from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DEFAULT_FILE_NAME = "report.xlsx"
MAX_SHEET_NAME_LENGTH = 31
MAX_COLUMN_WIDTH = 50
INVALID_SHEET_CHARS_RE = re.compile(r"[:\\/?*\[\]]")
INVALID_FILE_CHARS_RE = re.compile(r'[<>:"/\\|?*]+')


def normalize_file_name(file_name: str | None) -> str:
    safe_name = (file_name or DEFAULT_FILE_NAME).strip()
    safe_name = INVALID_FILE_CHARS_RE.sub("_", safe_name).strip(" .")

    if not safe_name:
        safe_name = DEFAULT_FILE_NAME

    if not safe_name.lower().endswith(".xlsx"):
        safe_name = f"{safe_name}.xlsx"

    return safe_name


def normalize_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS_RE.sub(" ", name).strip().strip("'")
    cleaned = re.sub(r"\s+", " ", cleaned)
    base_name = (cleaned or "Sheet")[:MAX_SHEET_NAME_LENGTH]

    candidate = base_name
    counter = 2

    while candidate.lower() in used_names:
        suffix = f"_{counter}"
        trimmed = base_name[: MAX_SHEET_NAME_LENGTH - len(suffix)].rstrip()
        candidate = f"{trimmed}{suffix}"
        counter += 1

    used_names.add(candidate.lower())
    return candidate


def collect_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()

    for row in rows:
        for key in row.keys():
            if key not in seen:
                headers.append(key)
                seen.add(key)

    return headers


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, MAX_COLUMN_WIDTH)


def write_sheet(worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return

    headers = collect_headers(rows)
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def build_xlsx_bytes(*, sheets: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names: set[str] = set()

    for sheet in sheets:
        sheet_name = normalize_sheet_name(sheet["name"], used_names)
        worksheet = workbook.create_sheet(title=sheet_name)
        write_sheet(worksheet, sheet["rows"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
