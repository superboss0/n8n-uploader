#!/usr/bin/env python3
import sys
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from copy import copy

# жёлтая заливка для несовпадений
YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# соответствие валют → названия колонок
debit_cols  = {'EUR': 'Дебет €', 'USD': 'Дебет $', 'RUB': 'Дебет ₽'}
credit_cols = {'EUR': 'Кредит €','USD': 'Кредит $','RUB': 'Кредит ₽'}

def copy_row(src_row, dst_ws, dst_row_idx):
    for src_cell in src_row:
        dst = dst_ws.cell(row=dst_row_idx, column=src_cell.col_idx, value=src_cell.value)
        if src_cell.has_style:
            dst.font          = copy(src_cell.font)
            dst.border        = copy(src_cell.border)
            dst.fill          = copy(src_cell.fill)
            dst.number_format = copy(src_cell.number_format)
            dst.protection    = copy(src_cell.protection)
            dst.alignment     = copy(src_cell.alignment)

def get_column_index_map(header_row):
    return {cell.value: idx for idx, cell in enumerate(header_row, start=1)}

def remove_empty_columns(ws):
    to_delete = [
        col for col in range(1, ws.max_column+1)
        if all(ws.cell(r, col).value in (None, "") for r in range(2, ws.max_row+1))
    ]
    for col in reversed(to_delete):
        ws.delete_cols(col)

def sort_sheet(ws, debit_col_name, credit_col_name):
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    data = [[ws.cell(r, c).value for c in range(1, ws.max_column+1)]
            for r in range(2, ws.max_row+1)]
    def val(x):
        try: return float(x)
        except: return float('-inf')
    di = hdr.index(debit_col_name)
    ci = hdr.index(credit_col_name)
    data.sort(key=lambda row: (val(row[di]), val(row[ci])), reverse=True)
    for r, row_vals in enumerate(data, start=2):
        for c, v in enumerate(row_vals, start=1):
            ws.cell(r, c).value = v

def reconcile(in_path, out_path):
    src_wb = load_workbook(in_path)
    src_ws = src_wb.active
    hdr_map = get_column_index_map(next(src_ws.iter_rows()))

    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)

    # 1) создаём листы EUR, USD, RUB и копируем заголовок
    sheets = {}
    cursors = {}
    for cur in ('EUR','USD','RUB'):
        ws = dst_wb.create_sheet(cur)
        copy_row(src_ws[1], ws, 1)
        sheets[cur] = ws
        cursors[cur] = 2

    # 2) фильтрация по валютам
    for row in src_ws.iter_rows(min_row=2):
        def v(col_name):
            x = row[hdr_map[col_name]-1].value
            try: return float(x)
            except: return 0.0

        for cur in ('EUR','USD','RUB'):
            if v(debit_cols[cur]) != 0 or v(credit_cols[cur]) != 0:
                copy_row(row, sheets[cur], cursors[cur])
                cursors[cur] += 1

    # 3) удаляем пустые колонки и сортируем
    for cur, ws in sheets.items():
        remove_empty_columns(ws)
        sort_sheet(ws, debit_cols[cur], credit_cols[cur])

    # 4) сопоставление и подсветка
    for cur, ws in sheets.items():
        hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        d_idx = hdr.index(debit_cols[cur]) + 1
        c_idx = hdr.index(credit_cols[cur]) + 1

        debits, credits = [], []
        for r in range(2, ws.max_row+1):
            d = ws.cell(r,d_idx).value or 0
            c = ws.cell(r,c_idx).value or 0
            cells = list(ws[r])
            if d and not c:
                debits.append(cells)
            elif c and not d:
                credits.append(cells)

        used, pairs = set(), []
        for d_cells in debits:
            amt = d_cells[d_idx-1].value
            found = False
            for i, c_cells in enumerate(credits):
                if i not in used and c_cells[c_idx-1].value == amt:
                    pairs.append((d_cells, c_cells, True))
                    used.add(i)
                    found = True
                    break
            if not found:
                pairs.append((d_cells, None, False))
        for i, c_cells in enumerate(credits):
            if i not in used:
                pairs.append((None, c_cells, False))

        out = dst_wb.create_sheet(f'{cur}_matched')
        out.append([f'D_{h}' for h in hdr] + [f'C_{h}' for h in hdr])

        for d_cells, c_cells, ok in pairs:
            vals, styles = [], []
            for block in (d_cells, c_cells):
                if block:
                    vals += [cell.value for cell in block]
                    styles += list(block)
                else:
                    vals += [None]*len(hdr)
                    styles += [None]*len(hdr)
            r2 = out.max_row+1
            for j, (v_src, s) in enumerate(zip(vals, styles), start=1):
                cell = out.cell(r2, j, value=v_src)
                if s and s.has_style:
                    cell.font          = copy(s.font)
                    cell.border        = copy(s.border)
                    cell.number_format = copy(s.number_format)
                    cell.alignment     = copy(s.alignment)
                if not ok:
                    cell.fill = YELLOW

        # автоширина
        for col in out.columns:
            mx = max(len(str(c.value)) if c.value else 0 for c in col)
            out.column_dimensions[col[0].column_letter].width = mx + 2

    # 5) сохраняем
    dst_wb.save(out_path)

def main():
    if len(sys.argv) < 2:
        print("Usage: python reconcile.py <input_file.xlsx>")
        sys.exit(1)

    in_path = sys.argv[1]
    base, _ = os.path.splitext(os.path.basename(in_path))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_filename = f"file_processed_{timestamp}.xlsx"
    out_path = os.path.join(os.path.dirname(in_path) or '.', out_filename)

    reconcile(in_path, out_path)
    print(f"✅ Обработка завершена: {out_path}")

if __name__ == "__main__":
    import sys
    main()
