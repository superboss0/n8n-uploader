import datetime
import os
import tempfile
from typing import Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

SMALL_DIFF = 1.0
TOLERANCE = 100.0

MERCHANT_MAPPING = {}

FIELDS = [
    "Start_balance",
    "Payin",
    "Payin_commission",
    "Payout",
    "Payout_commission",
    "Deposit",
    "Cashout",
    "Transfer",
    "End_balance",
]

DELTA_COLS = [f"Δ_{field}" for field in FIELDS]


def normalize_name(name: str) -> str:
    if pd.isna(name):
        return ""
    return str(name).strip().upper()


def map_merchant(name: str) -> str:
    base = str(name).strip()
    return MERCHANT_MAPPING.get(base, base)


def detect_name_column(df: pd.DataFrame, candidates) -> str:
    cols = set(df.columns)
    for candidate in candidates:
        if candidate in cols:
            return candidate
    raise ValueError(f"Could not find any of {candidates} in {list(df.columns)}")


def detect_report_kind_from_df_a(df_a: pd.DataFrame) -> str:
    cols = set(df_a.columns)
    if "merchants_name" in cols or "merchant_name" in cols:
        return "merchants"
    if "operator_name" in cols or "operators_name" in cols:
        return "operators"
    return "mixed"


def pick_col(df: pd.DataFrame, candidates, default_value=0):
    for candidate in candidates:
        if candidate in df.columns:
            return df[candidate]
    return pd.Series([default_value] * len(df), index=df.index)


def style_details_sheet(path: str, sheet_name: str = "Details") -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                text = ""
            elif isinstance(value, (int, float)):
                text = f"{value:,.2f}"
            else:
                text = str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = max_len + 2

    header_to_col = {cell.value: cell.column for cell in ws[1] if cell.value}

    delta_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    for delta_name in DELTA_COLS:
        col_idx = header_to_col.get(delta_name)
        if not col_idx:
            continue
        for cell in ws[get_column_letter(col_idx)]:
            cell.fill = delta_fill

    balance_fill = PatternFill(start_color="FFECF1DF", end_color="FFECF1DF", fill_type="solid")
    for col_name in ["MB_Start_balance", "FG_Start_balance"]:
        col_idx = header_to_col.get(col_name)
        if not col_idx:
            continue
        for cell in ws[get_column_letter(col_idx)]:
            cell.fill = balance_fill

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    wb.save(path)


def reconcile_files(input_a_path: str, input_b_path: str) -> Tuple[str, str]:
    df_a = pd.read_excel(input_a_path)
    df_a.rename(columns={col: col.lower() for col in df_a.columns}, inplace=True)
    report_kind = detect_report_kind_from_df_a(df_a)

    merchant_col_a = detect_name_column(
        df_a,
        ["merchants_name", "merchant_name", "operator_name", "operators_name", "name"],
    )

    df_a["Merchant_raw_A"] = df_a[merchant_col_a].astype(str)
    df_a["Merchant_mapped"] = df_a["Merchant_raw_A"].map(map_merchant)
    df_a["Merchant_key"] = df_a["Merchant_mapped"].map(normalize_name)
    df_a["Currency"] = df_a["currency"]

    df_a_norm = pd.DataFrame()
    df_a_norm["Merchant_key"] = df_a["Merchant_key"]
    df_a_norm["Merchant_A"] = df_a["Merchant_mapped"]
    df_a_norm["Currency"] = df_a["Currency"]
    df_a_norm["Start_balance_A"] = pick_col(df_a, ["start_balance"])
    df_a_norm["End_balance_A"] = pick_col(df_a, ["end_balance"])
    df_a_norm["Transfer_A"] = pick_col(df_a, ["transfer"])
    df_a_norm["Payin_A"] = pick_col(df_a, ["pay_in_turn", "payin", "pay_in", "payin_turnover"])
    df_a_norm["Payin_commission_A"] = pick_col(
        df_a, ["pay_in_commission", "payin_commission", "pay_in_fee", "payin_fee"]
    )
    df_a_norm["Payout_A"] = pick_col(df_a, ["payout_turn", "payout", "pay_out", "payout_turnover"])
    df_a_norm["Payout_commission_A"] = pick_col(
        df_a, ["payout_commission", "payout_fee", "pay_out_commission"]
    )
    df_a_norm["Deposit_A"] = pick_col(df_a, ["deposit", "deposit_turnover"])
    df_a_norm["Cashout_A"] = pick_col(df_a, ["cashout", "cash_out", "cashout_turnover"])
    df_a_norm["in_A"] = True

    df_b = pd.read_excel(input_b_path)
    df_b.rename(columns={col: col.lower() for col in df_b.columns}, inplace=True)
    df_b["Merchant_raw_B"] = df_b["name"].astype(str)
    df_b["Merchant_mapped"] = df_b["Merchant_raw_B"].map(map_merchant)
    df_b["Merchant_key"] = df_b["Merchant_mapped"].map(normalize_name)
    df_b["Currency"] = df_b["currency"]

    df_b_norm = pd.DataFrame()
    df_b_norm["Merchant_key"] = df_b["Merchant_key"]
    df_b_norm["Merchant_B"] = df_b["Merchant_mapped"]
    df_b_norm["Currency"] = df_b["Currency"]
    df_b_norm["Start_balance_B"] = df_b.get("start_balance", 0)
    df_b_norm["Payin_B"] = df_b.get("payin", 0)
    df_b_norm["Payin_commission_B"] = df_b.get("payin_commission", 0)
    df_b_norm["Payout_B"] = df_b.get("payout", 0)
    df_b_norm["Payout_commission_B"] = df_b.get("payout_commission", 0)
    df_b_norm["Deposit_B"] = df_b.get("deposit", 0)
    df_b_norm["Cashout_B"] = df_b.get("cashout", 0)
    df_b_norm["Transfer_B"] = df_b.get("transfer", 0)
    df_b_norm["End_balance_B"] = df_b.get("end_balance", 0)
    df_b_norm["in_B"] = True

    b_value_cols = [
        "Start_balance_B",
        "Payin_B",
        "Payin_commission_B",
        "Payout_B",
        "Payout_commission_B",
        "Deposit_B",
        "Cashout_B",
        "Transfer_B",
        "End_balance_B",
    ]
    multiplier = -1.0 if report_kind == "merchants" else 1.0
    for col in b_value_cols:
        df_b_norm[col] = multiplier * df_b_norm[col].fillna(0).astype(float)

    df = pd.merge(df_a_norm, df_b_norm, on=["Merchant_key", "Currency"], how="outer")
    df["in_A"] = df["in_A"].fillna(False).astype(bool)
    df["in_B"] = df["in_B"].fillna(False).astype(bool)
    df["Merchant"] = df["Merchant_A"].fillna(df["Merchant_B"])

    for field in FIELDS:
        for suffix in ["_A", "_B"]:
            col = field + suffix
            if col not in df.columns:
                df[col] = 0
            df[col] = df[col].fillna(0).astype(float)

    for field in FIELDS:
        df[f"Δ_{field}"] = df[f"{field}_A"] - df[f"{field}_B"]

    df["Max_abs_diff"] = df[DELTA_COLS].abs().max(axis=1)

    def calc_status(row):
        if row["in_A"] and not row["in_B"]:
            return "Только_в_A"
        if row["in_B"] and not row["in_A"]:
            return "Только_в_B"
        if row["Max_abs_diff"] <= SMALL_DIFF:
            return "OK"
        if row["Max_abs_diff"] <= TOLERANCE:
            return "Малая_разница"
        return "Ошибка"

    df["Status"] = df.apply(calc_status, axis=1)
    df["Comment"] = ""
    df.rename(columns={"Merchant": "Name"}, inplace=True)

    rename_map = {}
    for field in FIELDS:
        rename_map[f"{field}_A"] = f"MB_{field}"
        rename_map[f"{field}_B"] = f"FG_{field}"
    df.rename(columns=rename_map, inplace=True)

    ordered_cols = (
        ["Name", "Currency", "Status", "Max_abs_diff", "Comment"]
        + [f"MB_{field}" for field in FIELDS]
        + [f"FG_{field}" for field in FIELDS]
        + DELTA_COLS
        + ["Merchant_A", "Merchant_B", "Merchant_key", "in_A", "in_B"]
    )
    df = df[ordered_cols]

    status_order = ["Ошибка", "Малая_разница", "Только_в_A", "Только_в_B", "OK"]
    df["Status_order"] = df["Status"].apply(
        lambda status: status_order.index(status) if status in status_order else 999
    )
    df.sort_values(["Status_order", "Name", "Currency"], inplace=True)
    df.drop(columns=["Status_order"], inplace=True)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(tempfile.gettempdir(), f"reconcile_{report_kind}_{ts}.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Details", index=False)

    style_details_sheet(out_path)
    return out_path, report_kind
